import hashlib
import io
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from .model import NormalizedTransaction, Platform, RowStatus, amount_to_cents, normalize_note, normalize_occurred_at

WECHAT_SUCCESS = {
    "支付成功",
    "已到账",
    "对方已收钱",
    "已转账",
    "充值完成",
    "已存入零钱",
    "已转入零钱通",
    "提现已到账",
}
WECHAT_REFUND_PREFIXES = ("已全额退款", "已退款")

# 表头列（0 基）：交易时间,交易类型,交易对方,商品,收/支,金额(元),支付方式,当前状态,交易单号,商户单号,备注
COL_TIME = 0
COL_TXN_TYPE = 1
COL_COUNTERPARTY = 2
COL_ITEM = 3
COL_DIRECTION = 4
COL_AMOUNT = 5
COL_PAY_METHOD = 6
COL_STATUS = 7
COL_TXN_ID = 8
COL_MERCHANT_ID = 9
COL_NOTE = 10
EXPECTED_COLS = 11

DIRECTION_MAP = {
    "支出": "expense",
    "收入": "income",
    "/": "neutral",
}


def _classify_status(status_text: str) -> RowStatus:
    text = status_text.strip()
    if text in WECHAT_SUCCESS:
        return RowStatus.SUCCESS
    if text.startswith(WECHAT_REFUND_PREFIXES):
        return RowStatus.REFUND
    return RowStatus.SKIPPED


def _normalized_hash(
    platform: str,
    source_txn_id: str,
    occurred_at: str,
    amount_cents: int,
    direction: str,
    counterparty: str,
    item_desc: str,
) -> str:
    payload = "|".join(
        [
            platform,
            source_txn_id,
            occurred_at,
            str(amount_cents),
            direction,
            counterparty,
            item_desc,
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def parse_wechat_xlsx(path: str | Path) -> list[NormalizedTransaction]:
    """解析微信支付账单 XLSX，返回标准化行。

    前部为导出信息，表头行首个单元格为“交易时间”。
    损坏/不完整/伪造的 .xlsx（非 ZIP 字节、ZIP 内 XML 损坏、
    缺少 OOXML 必需部件、openpyxl 无法识别的工作簿）统一转换为
    ValueError，供导入路由渲染错误页，而不是暴露 HTTP 500。
    """
    try:
        return _parse_wechat_xlsx_workbook(path)
    except (zipfile.BadZipFile, InvalidFileException, ET.ParseError, KeyError) as exc:
        raise ValueError(f"微信账单文件损坏或无法解析：{exc}") from exc


def _parse_wechat_xlsx_workbook(path: str | Path) -> list[NormalizedTransaction]:
    # 从内存字节解析：openpyxl 解析失败时不会关闭其内部 ZipFile，
    # 直接读磁盘文件会在 Windows 上锁住临时文件，导致调用方无法清理；
    # 用 BytesIO 解析则磁盘临时文件从未被持有。
    workbook = openpyxl.load_workbook(
        io.BytesIO(Path(path).read_bytes()), read_only=True, data_only=True
    )
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header_row = None
        header_idx = -1
        data_rows = []
        for idx, row in enumerate(rows):
            if row and row[0] == "交易时间":
                header_idx = idx
                header_row = list(row)
                break
        if header_row is None:
            raise ValueError("微信账单未找到表头行（交易时间）")
        for row in rows:
            data_rows.append(list(row))
    finally:
        workbook.close()

    normalized: list[NormalizedTransaction] = []
    for row in data_rows:
        if not row or row[0] is None:
            continue
        cells = _align_cells(row)
        normalized.append(_normalize_row(cells))
    return normalized


def _align_cells(row: list) -> list:
    cells = list(row)
    if len(cells) < EXPECTED_COLS:
        cells += [None] * (EXPECTED_COLS - len(cells))
    return cells[:EXPECTED_COLS]


def _normalize_row(cells: list) -> NormalizedTransaction:
    occurred_at = normalize_occurred_at(cells[COL_TIME])
    raw_type = str(cells[COL_TXN_TYPE] or "").strip()
    counterparty = str(cells[COL_COUNTERPARTY] or "").strip()
    item_desc = str(cells[COL_ITEM] or "").strip()
    direction_raw = str(cells[COL_DIRECTION] or "").strip()
    amount_cents = amount_to_cents(cells[COL_AMOUNT])
    status_text = str(cells[COL_STATUS] or "").strip()
    source_txn_id = str(cells[COL_TXN_ID] or "").strip()
    note = normalize_note(cells[COL_NOTE])

    direction = DIRECTION_MAP.get(direction_raw, "neutral")
    status = _classify_status(status_text)
    normalized_hash = _normalized_hash(
        Platform.WECHAT.value,
        source_txn_id,
        occurred_at,
        amount_cents,
        direction,
        counterparty,
        item_desc,
    )
    return NormalizedTransaction(
        platform=Platform.WECHAT.value,
        source_txn_id=source_txn_id,
        occurred_at=occurred_at,
        amount_cents=amount_cents,
        direction=direction,
        status=status,
        status_text=status_text,
        counterparty=counterparty,
        item_desc=item_desc,
        raw_type=raw_type,
        note=note,
        normalized_hash=normalized_hash,
    )
